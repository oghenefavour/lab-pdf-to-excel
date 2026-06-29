"""
Lab PDF -> Excel  (Desktop App)
===================================================================
Rewritten to parse the specific lab report format shown in:
"Laboratory Report AN 10299.pdf"

Extracts:
  - Barcode -> Tube ID
  - Birthdate
  - Gender -> Sex (M/F)
  - Collection date
  - Physician (cleaned)
  - Tumour markers (AFP, CA125, CA19-9, CA15-3, CEA, CYFRA21-1)

Other fields (Patient ID, physician affiliation, email, symptoms) are left blank.
===================================================================
"""

import os
import re
import sys
import datetime
import shutil

try:
    import pdfplumber
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError as e:
    import tkinter as tk
    from tkinter import messagebox
    r = tk.Tk()
    r.withdraw()
    messagebox.showerror(
        "Missing library",
        f"{e}\n\nOpen a terminal and run:\n\npip install pdfplumber openpyxl"
    )
    sys.exit(1)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# -------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------
DATA_SHEET = "data"
FIRST_DATA_ROW = 14          # row 13 holds template headers/example

# field key -> (column letter, label shown in the form)
FIELDS = [
    ("tube_id",          "B", "Tube ID (Barcode)"),
    ("patient_id",       "C", "Patient ID (not in PDF)"),
    ("birth_date",       "D", "Birth Date (DD/MM/YYYY)"),
    ("sex",              "E", "Biological Sex"),
    ("collection_date",  "F", "Collection Date (DD/MM/YYYY)"),
    ("physician_name",   "G", "Ordering Physician Name"),
    ("physician_affil",  "H", "Physician Affiliation (not in PDF)"),
    ("email",            "I", "Reporting Email (not in PDF)"),
    ("language",         "J", "Language Code (default EN)"),
    ("symptoms",         "K", "Symptoms (not in PDF)"),
    ("AFP",              "L", "AFP (IU/mL)"),
    ("CA125",            "M", "CA125 (U/mL)"),
    ("CA15-3",           "N", "CA15-3 (U/mL)"),
    ("CA19-9",           "O", "CA19-9 (U/mL)"),
    ("CEA",              "P", "CEA (ng/mL)"),
    ("CYFRA21-1",        "Q", "CYFRA21-1 (ng/mL)"),
]
COLS = {k: c for k, c, _ in FIELDS}
DATE_FIELDS = {"birth_date", "collection_date"}
MARKER_FIELDS = {"AFP", "CA125", "CA15-3", "CA19-9", "CEA", "CYFRA21-1"}

# Patterns and units matching your PDF exactly (fixed for CA15-3 and CA19-9)
MARKERS = {
    "AFP":       (r"Alpha[- ]?fetoprotein",            r"IU/mL"),
    "CA125":     (r"Cancer\s+Antigen-?125",            r"U/mL"),
    "CA15-3":    (r"Cancer\s+Antigen[-.]?15[-.]?3",    r"U/mL"),   # handles hyphen or dot
    "CA19-9":    (r"Cancer\s+Antigen[-.]?19[-.]?9",    r"U/mL"),   # handles hyphen or dot
    "CEA":       (r"Carcinoembryonic\s+Antigen",       r"ng/mL"),
    "CYFRA21-1": (r"CYTOKERATIN\s+FRAGMENT\s+21-?1",   r"ng/mL"),
}

YELLOW = PatternFill("solid", fgColor="FFF2CC")

# -------------------------------------------------------------------
# PDF PARSING (rewritten for your report)
# -------------------------------------------------------------------
def read_pdf_text(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)

def _grab(pattern, text, group=1):
    m = re.search(pattern, text, re.IGNORECASE)
    return m.group(group).strip() if m else None

def _find_marker(text, name_pattern, unit_pattern):
    """Find a marker value that appears after the name and before the unit."""
    name_match = re.search(name_pattern, text, re.IGNORECASE)
    if not name_match:
        return None
    # Look ahead up to 150 characters for a number followed by the unit
    window = text[name_match.end(): name_match.end() + 150]
    # Allow decimal numbers, optional spaces, then the unit
    value_match = re.search(r"(\d+(?:\.\d+)?)\s*" + unit_pattern, window, re.IGNORECASE)
    return value_match.group(1) if value_match else None

def parse_pdf(pdf_path):
    """Return a dict of field -> string value (or '' if not found)."""
    text = read_pdf_text(pdf_path)
    d = {k: "" for k, _, _ in FIELDS}

    # Tube ID from "Barcode:"
    d["tube_id"] = _grab(r"Barcode\s*:\s*(\d+)", text) or ""

    # Birth date (DD/MM/YYYY)
    d["birth_date"] = _grab(r"Birthdate\s*:\s*(\d{2}/\d{2}/\d{4})", text) or ""

    # Gender: M -> Male, F -> Female
    gender = _grab(r"Gender\s*:\s*([MF])", text) or ""
    if gender == "M":
        d["sex"] = "Male"
    elif gender == "F":
        d["sex"] = "Female"
    else:
        d["sex"] = ""

    # Collection date
    d["collection_date"] = _grab(r"Collection date\s*:\s*(\d{2}/\d{2}/\d{4})", text) or ""

    # Physician name (clean trailing " Request date...")
    phys_raw = _grab(r"Physician\s*:\s*(.+?)(?:\n|$)", text) or ""
    d["physician_name"] = re.split(r"\s+Request", phys_raw)[0].strip()

    # Tumour markers
    for marker, (name_pat, unit_pat) in MARKERS.items():
        d[marker] = _find_marker(text, name_pat, unit_pat) or ""

    # Default language
    d["language"] = "EN"

    # All other fields (patient_id, physician_affil, email, symptoms) stay empty
    return d

def _parse_date(s):
    """Convert DD/MM/YYYY, DD-MM-YYYY, or YYYY-MM-DD to datetime."""
    s = (s or "").strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

# -------------------------------------------------------------------
# EXCEL APPEND
# -------------------------------------------------------------------
def first_empty_row(ws):
    row = FIRST_DATA_ROW
    while ws[f"{COLS['tube_id']}{row}"].value not in (None, ""):
        row += 1
    return row

def append_record(excel_path, form_values):
    """form_values: field -> raw string. Returns (row, missing_list)."""
    wb = load_workbook(excel_path)
    ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else wb.active
    row = first_empty_row(ws)

    missing = []
    for field, col in COLS.items():
        raw = (form_values.get(field) or "").strip()
        cell = ws[f"{col}{row}"]
        if raw == "":
            cell.fill = YELLOW
            missing.append(field)
            continue
        if field in DATE_FIELDS:
            dt = _parse_date(raw)
            if dt is None:
                wb.close()
                raise ValueError(
                    f"'{field}' must be a date like 31/10/1980 (got '{raw}')."
                )
            cell.value = dt
        elif field in MARKER_FIELDS:
            try:
                cell.value = float(raw)
            except ValueError:
                wb.close()
                raise ValueError(f"'{field}' must be a number (got '{raw}').")
        else:
            cell.value = raw
    wb.save(excel_path)
    return row, missing

# -------------------------------------------------------------------
# DESKTOP APP
# -------------------------------------------------------------------
class App(ttk.Frame):
    def __init__(self, master):
        super().__init__(master, padding=12)
        self.pack(fill="both", expand=True)
        self.excel_path = tk.StringVar()
        self.vars = {}
        self._build()

    def _build(self):
        master = self.master
        master.title("Lab PDF → Excel (for your report format)")
        master.geometry("700x780")   # FIXED: was "700+780" causing TclError

        # --- Working Excel file ---
        top = ttk.LabelFrame(self, text="1. Excel file to build (records are appended here)", padding=10)
        top.pack(fill="x", pady=(0, 10))
        row = ttk.Frame(top)
        row.pack(fill="x")
        ttk.Entry(row, textvariable=self.excel_path).pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="Choose…", command=self.choose_excel).pack(side="left", padx=4)
        ttk.Button(row, text="New copy from template…", command=self.new_from_template).pack(side="left")

        # --- Load a PDF ---
        mid = ttk.LabelFrame(self, text="2. Load a lab PDF, then check / complete the fields", padding=10)
        mid.pack(fill="both", expand=True, pady=(0, 10))
        ttk.Button(mid, text="Load PDF…", command=self.load_pdf).pack(anchor="w", pady=(0, 8))
        ttk.Button(mid, text="Add several PDFs at once (blanks left highlighted in Excel)…",
                   command=self.add_multiple_pdfs).pack(anchor="w", pady=(0, 8))

        form = ttk.Frame(mid)
        form.pack(fill="both", expand=True)
        # Left column: patient & order details (excluding markers)
        left_keys = [f for f in FIELDS if f[0] not in MARKER_FIELDS]
        # Right column: markers only
        right_keys = [f for f in FIELDS if f[0] in MARKER_FIELDS]

        colL = ttk.Frame(form)
        colL.pack(side="left", fill="both", expand=True, padx=(0, 12))
        colR = ttk.Frame(form)
        colR.pack(side="left", fill="y")

        ttk.Label(colL, text="Patient & order details", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        for key, _, label in left_keys:
            self._field(colL, key, label)

        ttk.Label(colR, text="Tumour markers (extracted from PDF)", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        for key, _, label in right_keys:
            self._field(colR, key, label)

        # --- Add button ---
        ttk.Button(self, text="3.  Add this record to the Excel file",
                   command=self.add_record).pack(fill="x", pady=(0, 10))

        # --- Log ---
        logf = ttk.LabelFrame(self, text="Activity log", padding=6)
        logf.pack(fill="both", expand=True)
        self.log = tk.Text(logf, height=7, wrap="word", state="disabled",
                           bg="#1e1e1e", fg="#d4d4d4", font=("Consolas", 9))
        self.log.pack(fill="both", expand=True)

    def _field(self, parent, key, label):
        f = ttk.Frame(parent)
        f.pack(fill="x", pady=2)
        ttk.Label(f, text=label, width=26, anchor="w").pack(side="left")
        var = tk.StringVar()
        e = tk.Entry(f, textvariable=var)
        e.pack(side="left", fill="x", expand=True)
        self.vars[key] = (var, e)

    # ---- actions ----
    def choose_excel(self):
        p = filedialog.askopenfilename(title="Choose the Excel file to append to",
                                       filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.excel_path.set(p)
            self._logline(f"Working file set: {p}")

    def new_from_template(self):
        src = filedialog.askopenfilename(title="Choose the blank template",
                                         filetypes=[("Excel files", "*.xlsx")])
        if not src:
            return
        dst = filedialog.asksaveasfilename(title="Save the new working file as…",
                                           defaultextension=".xlsx",
                                           filetypes=[("Excel files", "*.xlsx")])
        if not dst:
            return
        shutil.copyfile(src, dst)
        self.excel_path.set(dst)
        self._logline(f"New working file created: {dst}")

    def add_multiple_pdfs(self):
        path = self.excel_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("No Excel file",
                                   "Choose the Excel file to append to first (step 1).")
            return
        files = filedialog.askopenfilenames(
            title="Select one or more lab PDFs",
            filetypes=[("PDF files", "*.pdf")])
        if not files:
            return
        added = skipped = 0
        for f in files:
            try:
                data = parse_pdf(f)
                row, missing = append_record(path, data)
                pid = data.get("patient_id") or data.get("tube_id") or os.path.basename(f)
                note = f" (blank: {', '.join(missing)})" if missing else ""
                self._logline(f"Row {row}: added {pid} from {os.path.basename(f)}{note}")
                added += 1
            except PermissionError:
                messagebox.showerror("File is open",
                                     "Close the Excel file in Excel, then run the batch again.")
                break
            except Exception as ex:
                self._logline(f"SKIPPED {os.path.basename(f)}: {ex}")
                skipped += 1
        messagebox.showinfo(
            "Batch complete",
            f"Added {added} record(s)." +
            (f"\n{skipped} skipped (see log)." if skipped else "") +
            "\n\nOpen the Excel file and complete any yellow cells."
        )

    def load_pdf(self):
        p = filedialog.askopenfilename(title="Select the lab PDF",
                                       filetypes=[("PDF files", "*.pdf")])
        if not p:
            return
        try:
            data = parse_pdf(p)
        except Exception as ex:
            messagebox.showerror("Could not read PDF", str(ex))
            return
        for key, (var, entry) in self.vars.items():
            var.set(data.get(key, ""))
            # Highlight empty fields in light yellow
            entry.configure(bg="#fff2cc" if not data.get(key) else "white")
        miss = [k for k in COLS if not data.get(k)]
        self._logline(f"Loaded {os.path.basename(p)} — "
                      f"complete the highlighted fields: {', '.join(miss) if miss else 'none'}")

    def add_record(self):
        path = self.excel_path.get().strip()
        if not path:
            messagebox.showwarning("No Excel file", "Choose the Excel file to append to first (step 1).")
            return
        if not os.path.exists(path):
            messagebox.showerror("File not found", f"Cannot find:\n{path}")
            return
        values = {k: v.get() for k, (v, _) in self.vars.items()}
        if not values.get("tube_id", "").strip():
            messagebox.showwarning("Tube ID required",
                                   "Load a PDF or enter a Tube ID before adding a record.")
            return
        try:
            row, missing = append_record(path, values)
        except ValueError as ve:
            messagebox.showerror("Check a field", str(ve))
            return
        except PermissionError:
            messagebox.showerror("File is open",
                                 "Close the Excel file in Excel, then try again.")
            return
        pid = values.get("patient_id") or values.get("tube_id")
        note = f" (blank: {', '.join(missing)})" if missing else " (all fields complete)"
        self._logline(f"Row {row}: added {pid}{note}")
        # Clear the form for the next PDF
        for key, (var, entry) in self.vars.items():
            var.set("")
            entry.configure(bg="white")
        messagebox.showinfo("Added", f"Record added to row {row}.\n\nFile saved:\n{os.path.basename(path)}")

    def _logline(self, text):
        self.log.configure(state="normal")
        stamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.log.insert("end", f"[{stamp}] {text}\n")
        self.log.see("end")
        self.log.configure(state="disabled")

def main():
    root = tk.Tk()
    try:
        ttk.Style().theme_use("clam")
    except tk.TclError:
        pass
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()